"""
파일명: excel_file.py
목적: Excel 파일 읽기/쓰기 실습

사전 준비:
pip install pandas openpyxl
"""

print("=" * 70)
print("Excel 파일 처리 실습".center(70))
print("=" * 70)

# 1. pandas로 Excel 생성
print("\n[1] pandas로 Excel 파일 생성")
print("-" * 70)

try:
    import pandas as pd

    # 직원 데이터
    employees_data = {
        "사번": ["E001", "E002", "E003", "E004", "E005"],
        "이름": ["김철수", "이영희", "박민수", "정지훈", "최민지"],
        "부서": ["개발", "기획", "개발", "마케팅", "기획"],
        "직급": ["사원", "대리", "사원", "과장", "대리"],
        "연봉": [3500, 4200, 3200, 5000, 4000]
    }

    df = pd.DataFrame(employees_data)

    # Excel로 저장
    df.to_excel("employees.xlsx", index=False, sheet_name="직원명단")

    print("✓ employees.xlsx 생성 완료")
    print(f"\n데이터 미리보기:")
    print(df)

except ImportError:
    print("⚠️  pandas 또는 openpyxl이 설치되지 않았습니다.")
    print("   설치: pip install pandas openpyxl")
    exit()

# 2. Excel 파일 읽기
print("\n[2] Excel 파일 읽기")
print("-" * 70)

df = pd.read_excel("employees.xlsx", sheet_name="직원명단")

print(f"전체 데이터:")
print(df)

print(f"\n기본 정보:")
print(f"  행 수: {len(df)}")
print(f"  열 수: {len(df.columns)}")
print(f"  열 이름: {list(df.columns)}")

# 3. 데이터 분석
print("\n[3] 데이터 분석")
print("-" * 70)

print("연봉 통계:")
print(f"  평균: {df['연봉'].mean():,.0f}만원")
print(f"  최대: {df['연봉'].max():,}만원")
print(f"  최소: {df['연봉'].min():,}만원")
print(f"  중앙값: {df['연봉'].median():,.0f}만원")

print("\n부서별 인원:")
print(df['부서'].value_counts())

print("\n부서별 평균 연봉:")
dept_avg = df.groupby('부서')['연봉'].mean()
for dept, avg in dept_avg.items():
    print(f"  {dept}: {avg:,.0f}만원")

# 4. 여러 시트로 저장
print("\n[4] 여러 시트로 저장")
print("-" * 70)

# 부서별로 데이터 분리
with pd.ExcelWriter("employees_by_dept.xlsx", engine="openpyxl") as writer:
    # 전체 데이터
    df.to_excel(writer, sheet_name="전체", index=False)

    # 부서별 시트
    for dept in df['부서'].unique():
        dept_df = df[df['부서'] == dept]
        dept_df.to_excel(writer, sheet_name=dept, index=False)

print("✓ employees_by_dept.xlsx 생성 완료")
print("  - 시트: 전체, 개발, 기획, 마케팅")

# 5. 판매 데이터 생성 및 분석
print("\n[5] 판매 데이터 생성 및 분석")
print("-" * 70)

sales_data = {
    "날짜": pd.date_range("2024-01-01", periods=10),
    "제품": ["노트북", "마우스", "키보드", "모니터", "노트북",
             "마우스", "키보드", "노트북", "모니터", "키보드"],
    "수량": [3, 50, 30, 5, 2, 40, 25, 4, 3, 20],
    "단가": [1200000, 30000, 89000, 350000, 1200000,
             30000, 89000, 1200000, 350000, 89000]
}

sales_df = pd.DataFrame(sales_data)
sales_df['매출액'] = sales_df['수량'] * sales_df['단가']

# 저장
sales_df.to_excel("sales.xlsx", index=False, sheet_name="판매내역")

print("✓ sales.xlsx 생성 완료")
print(f"\n판매 데이터:")
print(sales_df.head())

# 제품별 집계
print("\n제품별 판매 현황:")
product_summary = sales_df.groupby('제품').agg({
    '수량': 'sum',
    '매출액': 'sum'
}).sort_values('매출액', ascending=False)

print(product_summary)

# 6. openpyxl로 세밀한 제어
print("\n[6] openpyxl로 스타일 적용")
print("-" * 70)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 새 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "스타일 적용 예제"

# 헤더 작성
headers = ["이름", "부서", "연봉", "평가"]
ws.append(headers)

# 데이터 작성
data = [
    ["김철수", "개발", 3500, "우수"],
    ["이영희", "기획", 4200, "우수"],
    ["박민수", "개발", 3200, "양호"],
    ["정지훈", "마케팅", 5000, "우수"],
]

for row in data:
    ws.append(row)

# 헤더 스타일
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 테두리
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=len(data)+1, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border

# 열 너비 조정
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10

# 저장
wb.save("styled_report.xlsx")

print("✓ styled_report.xlsx 생성 완료 (스타일 적용)")

# 7. 기존 Excel 파일 수정
print("\n[7] 기존 Excel 파일 수정")
print("-" * 70)

# 파일 열기
wb = load_workbook("employees.xlsx")
ws = wb.active

# 새 열 추가 (보너스)
ws['F1'] = "보너스"

for row in range(2, ws.max_row + 1):
    salary = ws[f'E{row}'].value
    bonus = salary * 0.1  # 연봉의 10%
    ws[f'F{row}'] = bonus

# 저장
wb.save("employees_with_bonus.xlsx")

print("✓ employees_with_bonus.xlsx 생성 완료 (보너스 열 추가)")

# 8. 대량 데이터 처리
print("\n[8] 대량 데이터 처리")
print("-" * 70)

import random

# 1000명의 직원 데이터 생성
large_data = {
    "사번": [f"E{i:04d}" for i in range(1, 1001)],
    "이름": [f"직원{i}" for i in range(1, 1001)],
    "부서": [random.choice(["개발", "기획", "마케팅", "영업"]) for _ in range(1000)],
    "연봉": [random.randint(3000, 8000) for _ in range(1000)]
}

large_df = pd.DataFrame(large_data)

# Excel로 저장
large_df.to_excel("large_employees.xlsx", index=False)

print(f"✓ large_employees.xlsx 생성 완료 ({len(large_df)}명)")

# 부서별 통계
print("\n부서별 통계:")
dept_stats = large_df.groupby('부서')['연봉'].agg(['count', 'mean', 'min', 'max'])
print(dept_stats)

# 9. 여러 Excel 파일 병합
print("\n[9] Excel 파일 병합")
print("-" * 70)

# 월별 판매 데이터 생성
for month in range(1, 4):
    month_data = {
        "제품": ["노트북", "마우스", "키보드"],
        "판매량": [random.randint(10, 50) for _ in range(3)],
        "매출": [random.randint(1000000, 5000000) for _ in range(3)]
    }
    df = pd.DataFrame(month_data)
    df.to_excel(f"sales_{month}월.xlsx", index=False)

print("✓ 3개월 판매 데이터 생성 완료")

# 병합
all_sales = []
for month in range(1, 4):
    df = pd.read_excel(f"sales_{month}월.xlsx")
    df['월'] = month
    all_sales.append(df)

merged_df = pd.concat(all_sales, ignore_index=True)
merged_df.to_excel("sales_merged.xlsx", index=False)

print("✓ sales_merged.xlsx 생성 완료 (3개월 병합)")

# 10. 피벗 테이블
print("\n[10] 피벗 테이블 생성")
print("-" * 70)

# 피벗 테이블 생성
pivot = merged_df.pivot_table(
    values='매출',
    index='제품',
    columns='월',
    aggfunc='sum',
    fill_value=0
)

print("월별 제품 매출:")
print(pivot)

# Excel로 저장
pivot.to_excel("sales_pivot.xlsx")

print("\n✓ sales_pivot.xlsx 생성 완료 (피벗 테이블)")

# 11. 생성된 파일 목록
print("\n[11] 생성된 Excel 파일")
print("-" * 70)

import os

excel_files = [f for f in os.listdir(".") if f.endswith('.xlsx')]

print(f"{'파일명':<30} {'크기':<15}")
print("-" * 45)

for filename in excel_files:
    size = os.path.getsize(filename)
    print(f"{filename:<30} {size:>10,} bytes")

print(f"\n총 {len(excel_files)}개의 Excel 파일 생성됨")

print("\n" + "=" * 70)
print("Excel 파일 처리 완료".center(70))
print("=" * 70)

print("\n💡 Tip: Excel 파일을 직접 열어서 확인해보세요!")
print("💡 Tip: pandas는 Excel 처리의 표준 도구입니다!")
print("💡 Tip: openpyxl로 서식을 세밀하게 제어할 수 있습니다!")
